"""
Weekly Excel Report Automator — Portfolio Sample
Author: Ali Raza
Description:
    A scheduled job that:
        1. Pulls fresh data from a SQL database (or CSVs)
        2. Builds a multi-sheet Excel report with charts
        3. Applies house-style formatting and conditional formatting
        4. Emails the workbook to a distribution list
        5. Logs success/failure to Slack

    Real-world use case: replaces 4 hours of manual Excel work every Monday
    morning with a 90-second cron job.

Usage:
    python 37_Excel_Report_Automator.py --config config.yaml
"""

import argparse
import logging
import smtplib
from datetime import date, timedelta
from email.message import EmailMessage
from pathlib import Path
from typing import Iterable

import pandas as pd
import requests
import yaml
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils.dataframe import dataframe_to_rows


logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s %(levelname)-7s %(message)s")
log = logging.getLogger("report")

NAVY = "1F3A68"
GOLD = "C9A227"
WHITE = "FFFFFF"
THIN = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def load_config(path: Path) -> dict:
    with open(path) as f:
        return yaml.safe_load(f)


def fetch_data(cfg: dict) -> dict[str, pd.DataFrame]:
    """Replace this with your real data layer (SQLAlchemy, BigQuery, etc.)."""
    today = date.today()
    last_week = today - timedelta(days=7)
    log.info("Fetching data for %s → %s", last_week, today)

    # Stub example — read CSV(s) configured in YAML
    return {
        name: pd.read_csv(path)
        for name, path in cfg["sources"].items()
    }


def style_header_row(ws, row: int, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(bold=True, color=WHITE, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def write_sheet(ws, df: pd.DataFrame, title: str) -> None:
    ws.cell(row=1, column=1, value=title).font = Font(
        bold=True, size=16, color=NAVY
    )
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    style_header_row(ws, row=2, cols=df.shape[1])
    for row in ws.iter_rows(min_row=3, max_row=2 + len(df), max_col=df.shape[1]):
        for cell in row:
            cell.border = BORDER


def add_revenue_chart(ws, df: pd.DataFrame) -> None:
    chart = LineChart()
    chart.title = "Daily Revenue (last 30 days)"
    chart.height = 10
    chart.width = 20
    last_col = df.shape[1]
    data = Reference(ws, min_col=2, max_col=last_col,
                     min_row=2, max_row=2 + len(df))
    cats = Reference(ws, min_col=1, min_row=3, max_row=2 + len(df))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, f"A{4 + len(df)}")


def build_workbook(data: dict[str, pd.DataFrame], out_path: Path) -> Path:
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, df in data.items():
        ws = wb.create_sheet(sheet_name[:31])
        write_sheet(ws, df, sheet_name.replace("_", " ").title())
        if "revenue" in sheet_name.lower():
            add_revenue_chart(ws, df)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    log.info("Workbook saved → %s", out_path)
    return out_path


def send_email(cfg: dict, attachment: Path) -> None:
    msg = EmailMessage()
    msg["From"] = cfg["email"]["from"]
    msg["To"] = ", ".join(cfg["email"]["to"])
    msg["Subject"] = f"Weekly Report — {date.today():%d %b %Y}"
    msg.set_content("Hi team,\n\nAttached is this week's report.\n\nBest,\nAli")
    msg.add_attachment(attachment.read_bytes(),
                       maintype="application",
                       subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                       filename=attachment.name)
    with smtplib.SMTP(cfg["email"]["smtp_host"], cfg["email"]["smtp_port"]) as s:
        s.starttls()
        s.login(cfg["email"]["user"], cfg["email"]["pass"])
        s.send_message(msg)
    log.info("Email sent to %s", msg["To"])


def post_slack(cfg: dict, msg: str) -> None:
    if not cfg.get("slack", {}).get("webhook"):
        return
    try:
        requests.post(cfg["slack"]["webhook"], json={"text": msg}, timeout=8)
    except Exception as exc:
        log.warning("Slack post failed: %s", exc)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--config", type=Path, required=True)
    args = ap.parse_args()

    cfg = load_config(args.config)
    out_path = Path(cfg["output_path"])

    try:
        data = fetch_data(cfg)
        wb_path = build_workbook(data, out_path)
        send_email(cfg, wb_path)
        post_slack(cfg, f"✅ Weekly report shipped: {wb_path.name}")
    except Exception as exc:
        log.exception("Report job failed")
        post_slack(cfg, f"❌ Weekly report FAILED: {exc}")
        raise


if __name__ == "__main__":
    main()
